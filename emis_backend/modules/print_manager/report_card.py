from datetime import datetime

import django
import os
from isdayoff import DateType, ProdCalendar
import asyncio
import uuid

import locale

locale.setlocale(locale.LC_ALL, ('ru_RU', 'UTF-8'))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')


from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def get_month(number):
    months = {
        '01': 'Январь',
        '02': 'Февраль',
        '03': 'Март',
        '04': 'Апрель',
        '05': 'Май',
        '06': 'Июнь',
        '07': 'Июль',
        '08': 'Август',
        '09': 'Сентябрь',
        '10': 'Октябрь',
        '11': 'Ноябрь',
        '12': 'Декабрь'
    }
    return months[number]


def get_days_to_month(date):
    m_work_days = []

    try:
        loop = asyncio.get_running_loop()
    except Exception as e:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    async def _calendar():
        nonlocal m_work_days
        calendar = ProdCalendar(locale='ru')
        m_work_days = await calendar.month(date)

    loop.run_until_complete(_calendar())
    return len(m_work_days)


def get_work_day(start_date, end_date, start_col, style):
    personnel_work_days = {}
    m_work_days = {}
    result = {}
    try:
        loop = asyncio.get_running_loop()
    except Exception as e:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    async def work_day():
        nonlocal personnel_work_days
        nonlocal m_work_days

        calendar = ProdCalendar(locale='ru')
        personnel_work_days = await calendar.range_date(start_date, end_date)
        m_work_days = await calendar.month(start_date)

    if start_date and end_date:
        loop.run_until_complete(work_day())

        for day, day_type in m_work_days.items():
            row = {
                'border': style['border'],
                'font': style['font'],
                'alignment': style['alignment']
            }
            idx = start_col
            if day_type == DateType.WORKING:
                row.update({'value': 'P', 'fill': PatternFill("solid", fgColor="FFFFFF")})

            if day_type == DateType.WORKING and personnel_work_days.get(day) is None:
                row.update({'value': 'H', 'fill': PatternFill("solid", fgColor="cccccc")})

            if day_type == DateType.NOT_WORKING:
                row.update({'value': 'B', 'fill': PatternFill("solid", fgColor="cccccc")})

            result.update({
                idx: row
            })
            start_col += 1

    return result


class ReportCartPrint:
    table_row = 14
    template: str = 'modules/print_manager/tpl/report_card.xlsx'
    tmp_file: str = f'temp/report_card_{uuid.uuid4()}.xlsx'
    style = {
        'font': Font('Arial', color="000000", size=12),
        'border': Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        ),
        'alignment': Alignment(horizontal="center", vertical="center")
    }

    def __init__(self, queryset):
        self.doc_id = None
        self.fields = None
        self.handler = None
        self.document = self.get_handler()
        self.queryset = queryset

    def get_handler(self):
        self.handler = load_workbook(self.template)
        return self.handler.active

    def header(self):
        period = self.queryset.period
        month = period.split('.')[0]
        year = period.split('.')[1]
        count_days = get_days_to_month(datetime.strptime(f'01.{period}', '%d.%m.%Y'))

        for i in range(1, count_days+1):

            day = f'{i}'
            if len(f'{i}') == 1:
                day = f'0{i}'

            self.document.cell(row=13, column=i+4).value = f'{day}.{period}'

        self.document.cell(row=12, column=5).value = f'за {get_month(month)} {year} г.'
        self.document.cell(row=3, column=1).value = self.queryset.entity.directorate
        self.document.cell(row=10, column=38).value = datetime.strptime(
            f'01.{period}', '%d.%m.%Y'
        ).strftime('%d %B %Y г.')
        self.document.cell(row=10, column=41).value = datetime.strptime(
            f'{count_days}.{period}', '%d.%m.%Y'
        ).strftime('%d %B %Y г.')

    def get_rows(self, field):
        return {
                1: {
                    'value': field.number,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': Alignment(horizontal="left", vertical="center")
                },
                2: {
                    'value': field.personnel.name,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': Alignment(horizontal="left", vertical="center")
                },
                3: {
                    'value': field.personnel.position,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': Alignment(horizontal="left", vertical="center")
                },
                4: {
                    'value': field.salary,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                },
                36: {
                    'value': field.work_day,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                },
                37: {
                    'value': field.start_date.strftime('%d.%m.%Y'),
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                },
                38: {
                    'value': field.end_date.strftime('%d.%m.%Y'),
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                },
                39: {
                    'value': field.salary_work,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                },
                40: {
                    'value': field.payment,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment'],
                    'fill': PatternFill("solid", fgColor="FFFFFF") if field.payment.strip() and int(field.payment) >= 0 else PatternFill("solid", fgColor="cccccc")
                },
                41: {
                    'value': field.award,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment'],
                    'fill': PatternFill("solid", fgColor="FFFFFF") if field.award.strip() and int(field.award) >= 0 else PatternFill("solid", fgColor="cccccc")
                },
                42: {
                    'value': field.penalty,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment'],
                    'fill': PatternFill("solid", fgColor="FFFFFF") if field.penalty.strip() and int(field.penalty) >= 0 else PatternFill("solid", fgColor="cccccc")
                },
                43: {
                    'value': field.accrued,
                    'border': self.style['border'],
                    'font': self.style['font'],
                    'alignment': self.style['alignment']
                }
            }

    def get_addon_field(self, row, field):
        days = get_work_day(field.start_date, field.end_date, 5, self.style)
        row.update(days)
        return row

    def table(self):
        default_cell = self.table_row
        number = 1

        for field in self.fields:
            field.number = number
            row = self.get_rows(field)
            row = self.get_addon_field(row, field)
            for col, value in row.items():
                cell = self.document.cell(row=default_cell, column=col)
                for key, val in value.items():
                    setattr(cell, key, val)
            number += 1
            default_cell += 1
            self.document.insert_rows(default_cell)

    def build(self):
        self.fields: QuerySet[HistoryTabel] = self.queryset.historytabel_set.all().order_by('personnel__name')
        self.header()
        self.table()
        return self.save()

    def save(self):
        filename = self.tmp_file
        self.handler.save(filename)
        return filename


if __name__ == '__main__':
    django.setup()
    from django.db.models import QuerySet
    from modules.crm.models import HistoryTabel, ReportCard

    queryset = ReportCard.objects.get(id=1)
    report_card = ReportCartPrint(queryset)
    report_card.build()


